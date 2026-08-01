"""
Exportação de empresas para arquivo .xlsx via openpyxl.
"""
import os
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def exportar_excel(empresas, nome_arquivo=None):
    """Gera arquivo .xlsx com as empresas fornecidas. Retorna caminho absoluto."""
    if not nome_arquivo:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"prospecao_{ts}.xlsx"

    caminho = os.path.join(RAIZ, nome_arquivo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empresas Prospectadas"

    fill_cabecalho = PatternFill("solid", fgColor="2563EB")
    fill_sem_site  = PatternFill("solid", fgColor="D1FAE5")
    fill_tem_site  = PatternFill("solid", fgColor="FEE2E2")
    fill_enviado   = PatternFill("solid", fgColor="DBEAFE")
    fill_site_demo = PatternFill("solid", fgColor="FEF3C7")

    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    fonte_link      = Font(color="2563EB", underline="single")
    borda_fina = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    colunas  = ["Nome da Empresa", "Telefone", "Endereço", "Categoria",
                "Nota Google", "Avaliações", "Tem Site?", "Status CRM",
                "Mensagem Enviada?", "Link Site Demo"]
    larguras = [40, 22, 55, 28, 13, 13, 12, 16, 20, 50]

    for col, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.font      = fonte_cabecalho
        cel.fill      = fill_cabecalho
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border    = borda_fina

    ws.row_dimensions[1].height = 22

    for linha, emp in enumerate(empresas, start=2):
        tem_site  = bool(emp.get("tem_site"))
        enviado   = bool(emp.get("mensagem_enviada"))
        nota      = emp.get("nota")
        avs       = emp.get("avaliacoes") or 0
        preview   = emp.get("preview_url") or ""

        linha_dados = [
            emp.get("nome", ""),
            emp.get("telefone", ""),
            emp.get("endereco", ""),
            emp.get("descricao_google") or emp.get("categoria", ""),
            f"{nota:.1f} ⭐" if nota else "",
            avs,
            "Sim" if tem_site else "Não",
            emp.get("status", "novo"),
            "Sim" if enviado else "Não",
            preview,
        ]

        for col, valor in enumerate(linha_dados, start=1):
            cel = ws.cell(row=linha, column=col, value=valor)
            cel.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            cel.border    = borda_fina

            if col == 7:  # Tem Site?
                cel.fill = fill_tem_site if tem_site else fill_sem_site
            elif col == 9 and enviado:
                cel.fill = fill_enviado
            elif col == 10 and preview:  # Link Site Demo
                cel.font = fonte_link
                cel.fill = fill_site_demo

        ws.row_dimensions[linha].height = 18

    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    wb.save(caminho)
    logger.info("Excel gerado: %s (%d empresas)", caminho, len(empresas))
    return caminho
